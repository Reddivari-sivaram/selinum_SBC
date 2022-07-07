


import time
import os
from win32 import *
import win32com.client as win32

import datetime as d
from openpyxl import *
import openpyxl as op
from selenium import *
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
prxy = ['10.2.179.100',
        '10.2.179.199',
        '192.168.10.200',
        '192.168.10.201',
        '10.104.34.80',
        '10.104.34.81',
        '10.105.34.67',
        '10.105.34.68']
special = ['10.192.98.20']
noC = ['10.61.1.25']    



xpath = {'10.16.163.41':'//*[@id="node_96"]/a','10.16.163.42':'//*[@id="node_84"]/a','10.1.0.4':'//*[@id="node_84"]/a',
        '10.1.0.5':'//*[@id="node_84"]/a','10.14.135.22':'//*[@id="node_84"]/a','10.192.98.20':'//*[@id="node_84"]/a',
        '10.192.198.17':'//*[@id="node_86"]/a','10.224.96.120':'//*[@id="node_98"]/a','10.36.20.171':'//*[@id="node_96"]/a',
        '10.33.19.20':'//*[@id="node_96"]/a','10.34.48.120':'//*[@id="node_98"]/a','10.35.128.120':'//*[@id="node_98"]/a',
        '10.33.210.50':'//*[@id="node_96"]/a','10.35.160.120':'//*[@id="node_98"]/a','10.33.240.120':'//*[@id="node_97"]/a',
        '10.33.240.81':'//*[@id="node_84"]/a','10.34.80.120':'//*[@id="node_98"]/a','10.33.80.120':'//*[@id="node_98"]/a',
        '10.35.96.120':'//*[@id="node_97"]/a','10.35.192.120':'//*[@id="node_98"]/a','10.33.176.120':'//*[@id="node_98"]/a',
        '10.33.160.120':'//*[@id="node_98"]/a','10.34.208.120':'//*[@id="node_98"]/a','10.34.16.120':'//*[@id="node_97"]/a',
        '10.34.16.114':'//*[@id="node_84"]/a','10.34.112.120':'//*[@id="node_97"]/a','10.34.176.120':'//*[@id="node_97"]/a',
        '10.38.138.177':'//*[@id="node_84"]/a','10.34.144.120':'//*[@id="node_97"]/a','10.33.112.120':'//*[@id="node_97"]/a',
        '10.192.98.18':'//*[@id="node_98"]/a','10.192.98.20':'//*[@id="node_84"]/a','10.105.34.67':'//*[@id="node_84"]/a',
        '10.105.34.68':'//*[@id="node_84"]/a','10.192.200.18':'//*[@id="node_99"]/a','10.225.34.11':'//*[@id="node_99"]/a',
        '10.2.179.100':'//*[@id="node_84"]/a','10.2.179.199':'//*[@id="node_84"]/a','192.168.10.200':'//*[@id="node_84"]/a',
        '192.168.10.201':'//*[@id="node_84"]/a','10.104.34.80':'//*[@id="node_84"]/a','10.104.34.81':'//*[@id="node_84"]/a',
        '10.61.0.23':'//*[@id="node_84"]/a','10.34.144.120':'//*[@id="node_97"]/a','10.192.195.13':'//*[@id="node_99"]/a',
        '10.30.32.71':'//*[@id="node_98"]/a','10.224.34.50':'//*[@id="node_96"]/a'}

wb = op.load_workbook("D:/test/exp.xlsx")
sh = wb['SBC']

rowcount = sh.max_row
for i in range(2,rowcount+1):
    path = "D:\chromedriver.exe"
    ip = str(sh.cell(i,3).value)
    if ip in prxy:
        try:
            driver = webdriver.Chrome(path)
            site = f"https://{sh.cell(i,3).value}/"
            driver.get(site)
            time.sleep(5)
            el = driver.find_element_by_id("details-button")
            time.sleep(3)
            el.click()
            el = driver.find_element_by_id("proceed-link")
            time.sleep(3)
            el.click()
            el1 = driver.find_element_by_name("splashbutton")
            time.sleep(3)
            el1.click()
            us = driver.find_element_by_id("username")
            ps = driver.find_element_by_id("password")
            us.send_keys("hcladmin")
            ps.send_keys("S0nU$5bc2)2!@hcl@0@1")
            el = driver.find_element_by_id('AckPreLoginMsg')
            time.sleep(2)
            el.click()
            el = driver.find_element_by_id('loginbutton')
            time.sleep(2)
            el.click()
            time.sleep(5)
            el = driver.find_element_by_id('tabSystem')
            time.sleep(0)
            el.click()
            time.sleep(20)
            el = driver.find_element_by_id('System_0_rt_Software_Base_Version-Span')
            sv = el.get_attribute("innerHTML").strip("'<b>'</b>''")
            el = driver.find_element_by_id('System_0_sysUptime-Span')
            ut = el.get_attribute("innerHTML").strip("'<b>'</b>''") 
            time.sleep(3)
            sh.cell(row=i,column=7,value =ut)
            sh.cell(row=i,column=9,value =sv)
            el = driver.find_element_by_id('tabSettings')
            el.click()
            time.sleep(21)
            el = driver.find_element_by_xpath(xpath[str(sh.cell(i,3).value)])
            el.click()
            time.sleep(9)
            el = driver.find_element_by_xpath('//*[@id="settingsCenterPane"]/div/div/ul/li[3]/ul[2]/li/div/table/tbody/tr/td[2]/a/b')
            el.click()
            time.sleep(9)
            el = driver.find_element_by_id('Settings_0_CertEndDateEpoch-Span')
            ced = el.get_attribute("innerHTML").strip("'<b>'</b>''")    # cer exp date
            sh.cell(row=i,column=11,value =ced)
            driver.close()
        except:
            driver.close()
            
    elif ip in noC:
        pass
    elif ip in special:
        try:
            driver = webdriver.Chrome(path)
            site = f"https://{sh.cell(i,3).value}/"
            driver.get(site)
            time.sleep(5)
            el = driver.find_element_by_id("details-button")
            time.sleep(3)
            el.click()
            el = driver.find_element_by_id("proceed-link")
            time.sleep(3)
            el.click()
            el1 = driver.find_element_by_name("splashbutton")
            time.sleep(3)
            el1.click()
            us = driver.find_element_by_id("username")
            ps = driver.find_element_by_id("password")
            us.send_keys("admin")
            ps.send_keys("S0nU$5bc1")
            el = driver.find_element_by_id('loginbutton')
            time.sleep(2)
            el.click()
            time.sleep(5)
            el = driver.find_element_by_xpath('//*[@id="ModalDiv"]/div/table/tbody/tr[2]/td/input')
            el.click()
            el = driver.find_element_by_id('tabSystem')
            time.sleep(5)
            el.click()
            time.sleep(20)
            el = driver.find_element_by_id('System_0_rt_Software_Base_Version-Span')
            sv = el.get_attribute("innerHTML").strip("'<b>'</b>''")
            el = driver.find_element_by_id('System_0_sysUptime-Span')
            ut = el.get_attribute("innerHTML").strip("'<b>'</b>''") 
            time.sleep(3)
            sh.cell(row=i,column=7,value =ut)
            sh.cell(row=i,column=9,value =sv)
            el = driver.find_element_by_id('tabSettings')
            el.click()
            time.sleep(21)
            el = driver.find_element_by_xpath(xpath[str(sh.cell(i,3).value)])
            el.click()
            time.sleep(9)
            el = driver.find_element_by_xpath('//*[@id="settingsCenterPane"]/div/div/ul/li[3]/ul[2]/li/div/table/tbody/tr/td[2]/a/b')
            el.click()
            time.sleep(9)
            el = driver.find_element_by_id('Settings_0_CertEndDateEpoch-Span')
            ced = el.get_attribute("innerHTML").strip("'<b>'</b>''")    # cer exp date
            sh.cell(row=i,column=11,value =ced)
            driver.close()
        except:
            driver.close()
    else:
        try:
            driver = webdriver.Chrome(path)
            site = f"https://{sh.cell(i,3).value}/"
            driver.get(site)
            time.sleep(5)
            el = driver.find_element_by_id("details-button")
            time.sleep(3)
            el.click()
            el = driver.find_element_by_id("proceed-link")
            time.sleep(3)
            el.click()
            el1 = driver.find_element_by_name("splashbutton")
            time.sleep(3)
            el1.click()
            us = driver.find_element_by_id("username")
            ps = driver.find_element_by_id("password")
            us.send_keys("admin")
            ps.send_keys("S0nU$5bc1")
            el = driver.find_element_by_id('loginbutton')
            time.sleep(2)
            el.click()
            time.sleep(5)
            el = driver.find_element_by_id('tabSystem')
            time.sleep(2)
            el.click()
            time.sleep(20)
            el = driver.find_element_by_id('System_0_rt_Software_Base_Version-Span')
            sv = el.get_attribute("innerHTML").strip("'<b>'</b>''")
            el = driver.find_element_by_id('System_0_sysUptime-Span')
            ut = el.get_attribute("innerHTML").strip("'<b>'</b>''") 
            time.sleep(3)
            sh.cell(row=i,column=7,value =ut)
            sh.cell(row=i,column=9,value =sv)
            el = driver.find_element_by_id('tabSettings')
            el.click()
            time.sleep(21)
            el = driver.find_element_by_xpath(xpath[str(sh.cell(i,3).value)])
            el.click()
            time.sleep(9)
            el = driver.find_element_by_xpath('//*[@id="settingsCenterPane"]/div/div/ul/li[3]/ul[2]/li/div/table/tbody/tr/td[2]/a/b')
            el.click()
            time.sleep(9)
            el = driver.find_element_by_id('Settings_0_CertEndDateEpoch-Span')
            ced = el.get_attribute("innerHTML").strip("'<b>'</b>''")    # cer exp date
            sh.cell(row=i,column=11,value =ced)
            driver.close()
        except:
            driver.close()
          


dt = str(d.datetime.today())[0:10]
fname = f"{dt}_SBC_Health_Report.xlsx"
sub = f"TEST2{dt}_SBC_Health_Report(Automated)"
loc = f"D:\\test\\{fname}"
wb.save(loc)
time.sleep(2)

olapp = win32.Dispatch('Outlook.Application')
olns = olapp.GetNameSpace('MAPI')
mail = olapp.CreateItem(0)
mail.display()
mail.Subject= sub
mail.BodyFormat =1
mail.Body= 'Hi Team,\n\nPlease find the SBC Health Report.\n\nNote: This mail is generated by automated scripts(python) it may miss the details of 1 or 2 sbc based on responce time of sbc while the scipt is calling.\n\nThanks,\nSiva'
mail.To = 'DL-HCLSkypeAdmin@biogen.com'
loc1 = f"D:/test/{fname}"
mail.Attachments.Add(os.path.join(os.getcwd(),loc1))
mail.save()
mail.send()

